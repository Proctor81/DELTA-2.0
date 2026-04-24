"""
DELTA - data/excel_export.py
Esportazione dei record di diagnosi in formato Excel (.xlsx).
Usa openpyxl con formattazione professionale per reportistica agronomica.
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional

from core.config import EXPORTS_DIR

logger = logging.getLogger("delta.data.excel")

EXPORT_PATH = Path(EXPORTS_DIR) / "delta_diagnoses.xlsx"

# Intestazioni colonne
HEADERS = [
    "Timestamp",
    "Temperatura (°C)",
    "Umidità (%)",
    "Pressione (hPa)",
    "Luminosità (lux)",
    "CO₂ (ppm)",
    "pH",
    "EC (mS/cm)",
    "Fonte sensori",
    "Classe AI",
    "Confidenza AI (%)",
    "Simulato",
    "Stato pianta",
    "Rischio globale",
    "Revisione umana",
    "Sintesi diagnosi",
    "Regole attivate",
    "Raccomandazioni",
]


class ExcelExporter:
    """
    Gestisce l'esportazione incrementale dei record DELTA su file Excel.
    Crea il file alla prima esportazione e aggiunge righe in append.
    """

    def __init__(self, export_path: Optional[Path] = None):
        self._path = export_path or EXPORT_PATH
        self._path.parent.mkdir(parents=True, exist_ok=True)

    # ─────────────────────────────────────────────
    # APPEND SINGOLO RECORD
    # ─────────────────────────────────────────────

    def append_record(self, record: Dict[str, Any]) -> bool:
        """
        Aggiunge un record al file Excel, creandolo se non esiste.

        Args:
            record: dict completo del record (da DeltaAgent.run_diagnosis)

        Returns:
            True se operazione riuscita
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("openpyxl non installato. Export Excel non disponibile.")
            return False

        try:
            if self._path.exists():
                wb = openpyxl.load_workbook(str(self._path))
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Diagnosi DELTA"
                self._write_header(ws)
                self._style_header(ws)

            row_data = self._record_to_row(record)
            ws.append(row_data)

            # Formatta riga appena aggiunta
            last_row = ws.max_row
            self._style_data_row(ws, last_row)

            # Adatta larghezza colonne
            self._auto_fit_columns(ws)

            wb.save(str(self._path))
            logger.debug("Record aggiunto all'Excel: %s (riga %d).", self._path, last_row)
            return True

        except Exception as exc:
            logger.error("Errore export Excel: %s", exc, exc_info=True)
            return False

    # ─────────────────────────────────────────────
    # EXPORT COMPLETO DA LISTA
    # ─────────────────────────────────────────────

    def export_all(self, records: List[Dict[str, Any]], path: Optional[Path] = None) -> bool:
        """
        Esporta una lista di record in un nuovo file Excel.

        Args:
            records: lista di record da esportare
            path:    percorso opzionale (default: path configurato)

        Returns:
            True se operazione riuscita
        """
        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl non installato.")
            return False

        target = path or self._path
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Diagnosi DELTA"
            self._write_header(ws)
            self._style_header(ws)

            for rec in records:
                ws.append(self._record_to_row(rec))

            for row_idx in range(2, ws.max_row + 1):
                self._style_data_row(ws, row_idx)

            self._auto_fit_columns(ws)
            wb.save(str(target))
            logger.info("Export Excel completato: %d record → %s.", len(records), target)
            return True
        except Exception as exc:
            logger.error("Errore export Excel batch: %s", exc, exc_info=True)
            return False

    # ─────────────────────────────────────────────
    # CONVERSIONE RECORD → RIGA
    # ─────────────────────────────────────────────

    @staticmethod
    def _record_to_row(record: Dict[str, Any]) -> List:
        """Converte un record dict in una lista di valori per Excel."""
        sd = record.get("sensor_data", {})
        ai = record.get("ai_result", {})
        dx = record.get("diagnosis", {})
        recs = record.get("recommendations", [])
        snap = dx.get("sensor_snapshot", sd)

        # Regole attivate → testo breve
        rules = dx.get("activated_rules", [])
        rules_text = "; ".join(
            f"[{r.get('risk','?').upper()}] {r.get('rule_id','?')}" for r in rules
        ) if rules else "Nessuna"

        # Raccomandazioni → testo multi-riga
        recs_text = "\n".join(
            f"[{r.get('category','?').upper()}] {r.get('action','')}" for r in recs
        ) if recs else "Nessuna"

        return [
            str(record.get("timestamp") or ""),
            snap.get("temperature_c"),
            snap.get("humidity_pct"),
            snap.get("pressure_hpa"),
            snap.get("light_lux"),
            snap.get("co2_ppm"),
            snap.get("ph"),
            snap.get("ec_ms_cm"),
            str(snap.get("source") or ""),
            str(ai.get("class") or ""),
            round(float(ai.get("confidence") or 0.0) * 100, 1),
            "Sì" if ai.get("simulated") else "No",
            str(dx.get("plant_status") or ""),
            str(dx.get("overall_risk") or ""),
            "Sì" if dx.get("needs_human_review") else "No",
            str(dx.get("summary") or ""),
            rules_text,
            recs_text,
        ]

    # ─────────────────────────────────────────────
    # STILE
    # ─────────────────────────────────────────────

    @staticmethod
    def _write_header(ws):
        """Scrive la riga di intestazione."""
        ws.append(HEADERS)

    @staticmethod
    def _style_header(ws):
        """Applica stile professionale all'intestazione."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 30
        except Exception:
            pass  # Stile non critico

    @staticmethod
    def _style_data_row(ws, row_idx: int):
        """Applica stile alle righe dati (alternanza colori)."""
        try:
            from openpyxl.styles import PatternFill, Alignment
            fill_color = "EBF3FB" if row_idx % 2 == 0 else "FFFFFF"
            fill = PatternFill("solid", fgColor=fill_color)
            for cell in ws[row_idx]:
                cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        except Exception:
            pass

    @staticmethod
    def _auto_fit_columns(ws):
        """Adatta la larghezza colonne al contenuto (con limite massimo)."""
        MAX_WIDTH = 60
        MIN_WIDTH = 12
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    val_str = str(cell.value)
                    line_len = max(len(line) for line in val_str.split("\n"))
                    max_len = max(max_len, line_len)
            ws.column_dimensions[col_letter].width = max(
                MIN_WIDTH, min(max_len + 2, MAX_WIDTH)
            )
